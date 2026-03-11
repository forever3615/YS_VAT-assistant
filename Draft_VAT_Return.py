import pandas as pd
import openpyxl
from openpyxl.styles import Font
import os
import sys
import warnings
import tkinter.messagebox as msg

# 忽略警告
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


def resource_path(relative_path):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)


def smart_read_ledger(path):
    """
    改进：智能定位科目余额表表头
    """
    # 先预览前10行寻找表头位置
    df_head = pd.read_excel(path, header=None, nrows=10)
    header_idx = 0
    for i, row in df_head.iterrows():
        if row.astype(str).str.contains('科目代码').any():
            header_idx = i
            break

    df = pd.read_excel(path, skiprows=header_idx)

    # 强制对应你代码中的13列结构
    standard_cols = ['科目代码', '科目名称', '公司', '年初借', '年初贷', '期初借', '期初贷', '本期借', '本期贷',
                     '本累计借', '本累计贷', '期末借', '期末贷']
    # 截取前13列并更名
    df = df.iloc[:, :len(standard_cols)]
    df.columns = standard_cols

    # 清洗：去掉科目代码为空或非数字的杂行
    df['科目代码'] = df['科目代码'].astype(str).str.strip()
    df = df[df['科目代码'].str.contains(r'\d', na=False)]

    # 金额列转数值
    for col in df.columns[3:]:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    return df


def generate_tax_report():
    try:
        def find_f(kw):
            for f in os.listdir('.'):
                if kw in f and not f.startswith('~$'): return f
            return None

        file_ledger = find_f('科目余额表')
        file_template = find_f('模版')
        file_input = find_f('用途确认')
        file_sales = find_f('全量发票')

        if not file_ledger or not file_template:
            msg.showerror("错误", "当前目录下未找到'科目余额表'或'模版'文件！")
            return

        # 1. 加载数据
        df_ledger = smart_read_ledger(file_ledger)
        df_mapping = pd.read_excel(file_template, sheet_name='科目对照')

        # 加载发票明细用于后期存档
        df_sales_raw = pd.read_excel(file_sales) if file_sales else None
        # 进项明细（用途确认）
        df_input_raw = pd.read_excel(file_input, skiprows=2) if file_input else None

        # 2. 预处理用途确认表（进项清单）汇总

        VALID_RATES = [0.01, 0.03, 0.05, 0.06, 0.09, 0.13]

        def find_nearest_rate(rate):
            """找到最接近的标准税率"""
            if pd.isna(rate):
                return 0
            # 计算当前税率与所有标准税率的差值，返回差值绝对值最小的那个
            return min(VALID_RATES, key=lambda x: abs(x - rate))

        input_list_summary = {}

        if file_input:
            try:
                df_inv = pd.read_excel(file_input, sheet_name='发票', skiprows=2)

                # 剔除含有“合计”字样的行
                mask = df_inv.apply(lambda row: row.astype(str).str.contains('合计').any(), axis=1)
                df_inv = df_inv[~mask]

                # 清洗无效数据
                df_inv = df_inv.dropna(subset=['金额', '税额'])
                df_inv = df_inv[df_inv['金额'] != 0]

                # 2. 计算原始税率
                df_inv['raw_rate'] = (df_inv['税额'] / df_inv['金额'])

                # 3. 核心步骤：归一化税率
                # 将 0.059 或 0.061 统一修正为 0.06
                df_inv['standard_rate'] = df_inv['raw_rate'].apply(find_nearest_rate)

                # 4. 按修正后的标准税率汇总
                summary_series = df_inv.groupby('standard_rate')['有效抵扣税额'].sum()

                # 5. 格式化输出为百分比格式
                input_list_summary = {k: round(v, 2) for k, v in summary_series.to_dict().items()}

            except Exception as e:
                print(f"解析并归一化税率失败: {e}")

        # 获取公司名
        company_name = "未知公司"
        if file_sales:
            df_s_name = pd.read_excel(file_sales, nrows=5)
            if '销方名称' in df_s_name.columns:
                company_name = str(df_s_name['销方名称'].iloc[0]).strip()

        # 3. 填报核对表
        wb = openpyxl.load_workbook(file_template)
        ws = wb['核对表']
        noto_font = Font(name='Noto Sans SC', size=9)

        def get_ledger_val(category, rate=None):
            sub_map = df_mapping[df_mapping['类别'] == category]
            if rate is not None:
                sub_map = sub_map[sub_map['税率'].astype(str).str.contains(str(rate))]

            codes = []
            for _, row in sub_map.iterrows():
                codes.extend([str(row['科目1']).strip(), str(row['科目2']).strip()])
            codes = [c for c in codes if c not in ['nan', 'None', '']]

            # 进项取借方，销项取贷方
            col_name = '本期借' if category == '进项税额' else '本期贷'
            return round(df_ledger[df_ledger['科目代码'].isin(codes)][col_name].sum(), 2)

        # 填充销项 (5-7行)
        for r, row in {0.13: 5, 0.09: 6, 0.06: 7}.items():
            ledger_val = get_ledger_val('销项税额', r)
            ws.cell(row=row, column=4, value=ledger_val).font = noto_font

        # 填充进项 (9-14行)
        for r, row in {0.13: 9, 0.09: 10, 0.06: 11, 0.05: 12, 0.03: 13, 0.01: 14}.items():
            ledger_val = get_ledger_val('进项税额', r)
            ws.cell(row=row, column=4, value=ledger_val).font = noto_font

            # E列：直接放差额 (账面 - 清单)
            list_val = input_list_summary.get(r, 0)
            diff_val = round(list_val - ledger_val, 2)
            ws.cell(row=row, column=5, value=diff_val).font = noto_font

        # 填充简易征收（25-27行）
        for r, row in {0.05: 25, 0.03: 26, 0.01:27}.items():
            ledger_val = get_ledger_val('简易计税', r)
            ws.cell(row=row, column=4, value=ledger_val).font = noto_font

        # 15行通行费处理
        toll_ledger = get_ledger_val('进项税额', '通行费')
        ws.cell(row=15, column=4, value=toll_ledger).font = noto_font
        #ws.cell(row=15, column=5, value=toll_ledger).font = noto_font  # 通行费默认差额为全额（除非有对应清单）

        # 进项转出明细 D18-D22
        tm = df_mapping[df_mapping['类别'] == '进项税转出']
        for i in range(5):
            val = 0
            if i < len(tm):
                c = [str(tm.iloc[i]['科目1']).strip(), str(tm.iloc[i]['科目2']).strip()]
                val = df_ledger[df_ledger['科目代码'].isin([x for x in c if x != 'nan'])]['本期贷'].sum()
            ws.cell(row=18 + i, column=4, value=round(val, 2)).font = noto_font

        # 销售清单轧差填充 (J-O 列)
        if df_sales_raw is not None:
            df_s = df_sales_raw.copy()
            df_s['rate_num'] = df_s['税率'].astype(str).str.rstrip('%').astype(float)
            if df_s['rate_num'].max() > 1: df_s['rate_num'] /= 100

            row_map = {0.13: 5, 0.09: 6, 0.06: 7, 0.05: 25, 0.03: 26, 0.01: 27}
            for r, r_idx in row_map.items():
                sp = df_s[(df_s['rate_num'] == r) & (df_s['发票票种'].str.contains('专用发票'))]
                pp = df_s[(df_s['rate_num'] == r) & (~df_s['发票票种'].str.contains('专用发票'))]
                vals = [sp['金额'].sum(), sp['税额'].sum(), pp['金额'].sum(), pp['税额'].sum()]
                for i, v in enumerate(vals):
                    ws.cell(row=r_idx, column=10 + i, value=round(v, 2)).font = noto_font
                # 轧差
                ledger_tax = ws.cell(row=r_idx, column=4).value or 0
                un_tax = round(ledger_tax - vals[1] - vals[3], 2)
                ws.cell(row=r_idx, column=14, value=round(un_tax / r, 2) if r != 0 else 0).font = noto_font
                ws.cell(row=r_idx, column=15, value=un_tax).font = noto_font

        # 4. 保存结果
        output_name = f"增值税申报底稿_{company_name}.xlsx"
        wb.save(output_name)

        with pd.ExcelWriter(output_name, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_ledger.to_excel(writer, sheet_name='TB', index=False)
            if df_sales_raw is not None:
                df_sales_raw.to_excel(writer, sheet_name='开票清单', index=False)
            if df_input_raw is not None:
                df_input_raw.to_excel(writer, sheet_name='进项清单', index=False)

        msg.showinfo("成功", f"底稿生成成功：\n{output_name}")

    except Exception as e:
        msg.showerror("错误", f"程序运行发生异常：\n{str(e)}")


if __name__ == '__main__':
    generate_tax_report()