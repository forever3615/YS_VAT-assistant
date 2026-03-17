from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from openpyxl import load_workbook


def finalize_and_beautify(target_file):
    print("\n>>> 正在进行最后的 Sheet 整理与格式美化...")

    # 1. 定义 Sheet 的目标顺序和保留清单
    # 放在最前面的两个
    priority_sheets = ['销项测算（含中台）', '中台测算核对表']
    # 后面依次排列的保留 Sheet
    other_keep_sheets = [
        '科目余额表', '利润中心余额表', '物业收入计提表',
        '收款汇总表', '收入结转汇总表', '中台明细汇总',
        '中台账套汇总', '不动产租赁', '利润中心映射待确认'
    ]
    all_keep = priority_sheets + other_keep_sheets

    wb = load_workbook(target_file)

    # --- A. 处理隐藏和排序 ---
    # 获取当前 wb 中所有的 sheet 名
    current_sheets = wb.sheetnames

    # 隐藏不需要的 Sheet
    for name in current_sheets:
        if name not in all_keep:
            wb[name].sheet_state = 'hidden'

    # 重新排列顺序 (按 all_keep 的顺序，存在才排)
    # openpyxl 重新排列是通过 wb._sheets 列表操作
    new_order = []
    for name in all_keep:
        if name in current_sheets:
            new_order.append(wb[name])
    # 将原本不在 keep 清单里的（即隐藏的）加在最后
    for name in current_sheets:
        if name not in all_keep:
            new_order.append(wb[name])
    wb._sheets = new_order

    # --- B. 核心 Sheet 美化 (前两个) ---
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')  # 浅蓝色
    border_style = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))
    header_font = Font(bold=True)

    for sheet_name in priority_sheets:
        if sheet_name not in current_sheets:
            continue

        ws = wb[sheet_name]

        # 冻结首行
        ws.freeze_panes = 'A2'

        # 遍历所有单元格设置对齐和边框
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_style

                # 表头特殊格式
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font

        # 自动调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 40)  # 最大不超过40

    wb.save(target_file)
    print(">>> 报表整理美化完成！核心 Sheet 已置顶，其余过程 Sheet 已隐藏。")

# 在你的主程序 Combined_OutputVAT.py 的最后一步调用：
# finalize_and_beautify(target_file)