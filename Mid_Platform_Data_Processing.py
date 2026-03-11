import pandas as pd
import os
import warnings
from openpyxl.utils import get_column_letter

# 忽略样式警告
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


# ==========================================
# 1. 内部逻辑模块
# ==========================================

def _clean_single_df(file_path, header, fill_idx, tag, rename_map=None):
    """
    清洗单张业务表：
    1. 填充合并单元格
    2. 收款表专用逻辑：剔除结算方式为空 或 结算方式为“小计”的行
    3. 其他表逻辑：剔除包含小计/合计关键字的行
    """
    if not file_path or not os.path.exists(file_path):
        return None

    # 读取数据
    df = pd.read_excel(file_path, header=header)

    # 1. 填充合并单元格 (前 N 列)
    df.iloc[:, :fill_idx] = df.iloc[:, :fill_idx].ffill()

    # 2. 核心清洗逻辑
    if tag == "收款表":
        if '结算方式' in df.columns:
            # A. 剔除结算方式为空的行
            df = df.dropna(subset=['结算方式'])
            # B. 剔除结算方式中包含“小计”字样的行
            df = df[~df['结算方式'].astype(str).str.contains('小计')]
    else:
        # 计提和结转表：全表扫描“小计/合计”关键字剔除
        mask = df.astype(str).apply(lambda x: x.str.contains('小计|合计')).any(axis=1)
        df = df[~mask]

    # 3. 拍扁多级表头 (针对结转表)
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = [f"{c[0]}_{c[1]}" if 'Unnamed' not in str(c[1]) else c[0] for c in df.columns]

    # 4. 字段重命名
    if rename_map:
        df = df.rename(columns=rename_map)

    return df.dropna(how='all')


def _apply_financial_formulas(df):
    """注入 Excel 计算公式 (H-O列逻辑)"""
    if df.empty: return df
    df = df.reset_index(drop=True)
    col_map = {col: get_column_letter(df.columns.get_loc(col) + 1) for col in df.columns}

    for i in range(len(df)):
        r = i + 2
        formula_total = (
            f'=ROUND({col_map["权责收入合计数"]}{r}-{col_map["本月增加以前年度应收"]}{r}+'
            f'{col_map["本月减免以前年度"]}{r}+{col_map["当期实收金额"]}{r}-'
            f'{col_map["当期收款_以前月份"]}{r}-{col_map["当期收款_当期"]}{r}-'
            f'{col_map["以前月份预收款转入"]}{r}-{col_map["以前年度预收款转入"]}{r}, 2)'
        )
        df.loc[i, '增值税申报价税合计'] = formula_total
        p_letter = get_column_letter(df.columns.get_loc('增值税申报价税合计') + 1)
        df.loc[i, '增值税申报销售额'] = f'=ROUND({p_letter}{r}/(1+{col_map["税率"]}{r}), 2)'
        q_letter = get_column_letter(df.columns.get_loc('增值税申报销售额') + 1)
        df.loc[i, '税额'] = f'=ROUND({p_letter}{r}-{q_letter}{r}, 2)'
    return df


# ==========================================
# 2. 主模块接口
# ==========================================


def process_revenue_summary(target_main_file, collection_file, carryover_file, provision_file):
    keys_detail = ['区域名称', '城市公司', '项目名称', '账套编码', '账套名称', '收费科目']
    keys_account = ['区域名称', '城市公司', '账套编码', '账套名称', '收费科目']
    val_cols = ['权责收入合计数', '本月增加以前年度应收', '本月减免以前年度', '当期实收金额',
                '当期收款_以前月份', '当期收款_当期', '以前月份预收款转入', '以前年度预收款转入']

    print("\n>>> 开始执行业务报表模块...")

    # 1. 分别清洗
    df_jt = _clean_single_df(provision_file, 0, 6, "计提表", {'合计金额': '权责收入合计数'})
    df_sk = _clean_single_df(collection_file, 0, 7, "收款表", {'发生额': '当期实收金额', '账套名': '账套名称'})
    df_jz = _clean_single_df(carryover_file, [0, 1], 7, "结转表")

    # 2. 【核心修复】在合并前，先对每张表进行预汇总，确保维度唯一
    dfs_to_merge = []

    if df_jt is not None:
        # 计提表：按维度汇总金额列
        df_jt_grp = df_jt.groupby(keys_detail).agg({
            '税率': 'first',
            '权责收入合计数': 'sum',
            '本月增加以前年度应收': 'sum',
            '本月减免以前年度': 'sum'
        }).reset_index()
        dfs_to_merge.append(df_jt_grp)

    if df_sk is not None:
        # 收款表：按维度汇总实收金额（干掉因结算方式产生的多行）
        df_sk_grp = df_sk.groupby(keys_detail).agg({
            '当期实收金额': 'sum'
        }).reset_index()
        dfs_to_merge.append(df_sk_grp)

    if df_jz is not None:
        # 结转表：按维度汇总四项预收转入
        df_jz_grp = df_jz.groupby(keys_detail).agg({
            '当期收款_以前月份': 'sum',
            '当期收款_当期': 'sum',
            '以前月份预收款转入': 'sum',
            '以前年度预收款转入': 'sum'
        }).reset_index()
        dfs_to_merge.append(df_jz_grp)

    # 3. 执行合并与写入逻辑
    if not os.path.exists(target_main_file):
        pd.DataFrame().to_excel(target_main_file)

    with pd.ExcelWriter(target_main_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        # 写入清洗后的原表 (注意：原表保留明细，汇总页使用聚合后的数据)
        if df_jt is not None: df_jt.to_excel(writer, sheet_name='物业收入计提表', index=False)
        if df_sk is not None: df_sk.to_excel(writer, sheet_name='收款汇总表', index=False)
        if df_jz is not None: df_jz.to_excel(writer, sheet_name='收入结转汇总表', index=False)

        if dfs_to_merge:
            from functools import reduce
            # 此时 Merge 是 1对1 的，不会再产生重复广播
            merged = reduce(lambda left, right: pd.merge(left, right, on=keys_detail, how='outer'),
                            dfs_to_merge).fillna(0)

            # A. 明细汇总
            df_detail = _apply_financial_formulas(merged[keys_detail + ['税率'] + val_cols].copy())
            df_detail.to_excel(writer, sheet_name='中台明细汇总', index=False)

            # B. 账套汇总
            df_account = df_detail.groupby(keys_account).agg(
                {'税率': 'first', **{col: 'sum' for col in val_cols}}).reset_index()
            df_account = df_account[(df_account[val_cols] != 0).any(axis=1)]
            df_account = _apply_financial_formulas(df_account)  # 账套页重新应用公式
            df_account.to_excel(writer, sheet_name='中台账套汇总', index=False)

    print(f">>> 处理完成。")


# ==========================================
# 3. 测试模块
# ==========================================

if __name__ == "__main__":
    # 此处替换为你本地的文件路径进行单独测试
    process_revenue_summary("总表测试.xlsx", "收款汇总表_20260310_1611.xlsx", "收入结转汇总表_20260310_1612.xlsx",
                            "物业收入计提表_20260310_1610.xlsx")